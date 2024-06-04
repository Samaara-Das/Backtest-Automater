from openpyxl import load_workbook
from components.logger import setup_logger

logger = setup_logger(__name__)

class SettingsReader:
    def __init__(self, excel_path):
        self.excel_path = excel_path
        self.logger = logger

    def read_settings(self):
        """
        Reads an Excel file and converts each row into a dictionary, where the keys are the column headers
        and the values are the corresponding cell values in that row.

        Returns:
            list of dict: A list of dictionaries, each representing a row in the Excel sheet.

        Raises:
            FileNotFoundError: If the Excel file is not found.
            ValueError: If the Excel sheet is empty or improperly formatted.
        """
        try:
            self.logger.info(f"Attempting to load the workbook from {self.excel_path}")
            workbook = load_workbook(self.excel_path) # Load the workbook and select the active sheet
            sheet = workbook.active

            headers = [cell.value for cell in sheet[1]] # Get the column headers
            if not headers:
                raise ValueError("No headers found in the Excel sheet")
            
            settings_list = []  # Convert each row to a dictionary with column headers as keys and cell values as values
            for row in sheet.iter_rows(min_row=2, values_only=True):
                if any(row):  # Ensure the row is not entirely empty
                    settings = {headers[i]: row[i] for i in range(len(headers))}
                    settings_list.append(settings)

            if not settings_list:
                raise ValueError("No data rows found in the Excel sheet")

            self.logger.info("Successfully read the Excel file and converted it to a list of dictionaries")
            return settings_list

        except FileNotFoundError:
            self.logger.error(f"The file at {self.excel_path} was not found.")
            raise
        except ValueError as ve:
            self.logger.error(f"ValueError: {ve}")
            raise
        except Exception as e:
            self.logger.error(f"An unexpected error occurred: {e}")
            raise

