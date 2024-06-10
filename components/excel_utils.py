'''
This module contains functions for working with the Backtest Report Data Excel file.
'''

import os
import openpyxl
from openpyxl import Workbook
from components.logger import setup_logger, INFO

# Set up logger for this file
main_logger = setup_logger(__name__, INFO)

class ExcelUtil:
    def __init__(self, report_data_file_path):
        self.file_path = report_data_file_path

    def setup_excel_file(self, titles):
        """
        Ensures the `self.file_path` Excel file exists and has the required headers.
        
        Args:
        - titles (dict): A dictionary where keys are the titles and values are the corresponding xpaths.

        Returns:
        - str: The path to the Excel file.
        """
        try:
            if not os.path.exists(self.file_path):  # Create a new workbook and add headers if file does not exist
                wb = Workbook()
                ws = wb.active
                ws.title = "Backtest Report Data"

                headers = ["Source File"] + list(titles.keys())
                ws.append(headers)

                wb.save(self.file_path)
                main_logger.info(f"Created new Excel file: {self.file_path}")

            else: # If the file exists, ensure all headers are present
                wb = openpyxl.load_workbook(self.file_path)
                ws = wb.active

                existing_headers = [cell.value for cell in ws[1]]
                new_headers = ["Source File"] + list(titles.keys())

                # Add any missing headers to the existing workbook
                for header in new_headers:
                    if header not in existing_headers:
                        ws.cell(row=1, column=len(existing_headers) + 1, value=header)
                        existing_headers.append(header)

                wb.save(self.file_path)
                main_logger.info(f"Verified and updated headers in existing Excel file: {self.file_path}")

            return self.file_path
        except Exception as e:
            main_logger.error(f"Error setting up Excel file {self.file_path}: {e}")
            raise

    def add_data_to_excel(self, data):
        """
        Add data to the Excel file. If a row with the same 'Source File' value exists, update it;
        otherwise, append a new row.
        
        Args:
        - data (dict): A dictionary where keys are column headers and values are the data to be added.
        """
        try:
            wb = openpyxl.load_workbook(self.file_path)
            ws = wb.active

            source_file = data["Source File"]
            existing_row_idx = None

            # Check if the same "Source File" value already exists in the Excel file
            for row in ws.iter_rows(min_row=2, max_col=1):
                if row[0].value == source_file:
                    existing_row_idx = row[0].row
                    break

            if existing_row_idx: # If data for this file has already been written in the Excel file, just update it
                for col_num, header in enumerate(ws[1], start=1):
                    ws.cell(row=existing_row_idx, column=col_num, value=data.get(header.value, ""))
                main_logger.info(f"Updated data for {source_file} in Excel file.")
            else: # Or else, add a new row for this new file
                new_row = []
                for header in ws[1]:
                    new_row.append(data.get(header.value, ""))
                ws.append(new_row)
                main_logger.info(f"Appended data for {source_file} to Excel file.")

            wb.save(self.file_path)
        except Exception as e:
            main_logger.error(f"Error adding data to Excel file {self.file_path}: {e}")
            raise