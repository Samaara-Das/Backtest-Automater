from os import path
import psutil
import time
from datetime import datetime
from openpyxl import load_workbook
from pywinauto import Application
from pywinauto.keyboard import send_keys
import re
from pyperclip import copy, paste
import logger_setup

# Set up the logger
logger = logger_setup.setup_logger(__name__)

# Constants
MT4_EXE_PATH = r"C:\Program Files (x86)\Tradeview MetaTrader 4 Terminal\terminal.exe" # MetaTrader 4 terminal
ME_EXE_PATH = r"C:\Program Files (x86)\Tradeview MetaTrader 4 Terminal\metaeditor.exe" # MetaEditor 4
TIMEOUT = 20 
EXCEL_PATH = "D:\\Strategy Tester Settings.xlsx"

def ea_base_name(ea_name):
    """
    Convert the EA name by removing any leading directory names and the '.ex4' extension, keeping only the final name part.
    
    Parameters:
    ea_name (str): The EA name to be converted.
    
    Returns:
    str: The converted string.
    """
    try:
        converted_string = re.sub(r'.*\\(.*)\.ex4$', r'\1', ea_name) # Using regex to remove the directory path and .ex4 extension
        logger.info(f"Converted '{ea_name}' to '{converted_string}'")
        return converted_string
    except Exception as e:
        logger.error(f"Error converting string '{ea_name}': {e}")
        return None

def is_application_running(exe_path):
    """
    Check if the application specified by exe_path is already running.
    
    Args:
        exe_path (str): The path to the executable file.
        
    Returns:
        int: The process ID (pid) if running, otherwise returns None.
    """
    logger.info("Checking if the application is already running.")
    for proc in psutil.process_iter(['pid', 'name', 'exe']):
        try:
            if proc.info['exe'] and path.normpath(proc.info['exe']) == path.normpath(exe_path):
                logger.info(f"Application is already running with PID: {proc.info['pid']}")
                return proc.info['pid']
        except (psutil.NoSuchProcess, psutil.AccessDenied, psutil.ZombieProcess) as e:
            logger.error(f"Exception occurred while checking running processes: {e}")
    logger.info("Application is not running.")
    return None

def replace_value(original_text, new_value):
    """
    Replaces the value between '=' and ';' in `original_text` with `new_value`.

    Args:
        original_text (str): The original string containing the value to be replaced.
        new_value (str): The new value to insert between '=' and ';'.

    Returns:
        str: The modified string with the replaced value.

    Raises:
        ValueError: If the input string does not contain '=' and/or ';' or other errors occur.
    """
    try:
        start_index = original_text.find('=') + 1
        end_index = original_text.find(';')
        
        if start_index == 0 or end_index == -1:
            raise ValueError("The input string does not contain '=' and/or ';'")
        
        modified_string = original_text[:start_index] + new_value + original_text[end_index:]
        logger.info("Successfully replaced the value in the string.")
        
        return modified_string
    except Exception as e:
        logger.error(f"An error occurred in replace_value: {e}")
        raise

def change_input_value(text, prop, value):
    """
    Finds and replaces the value of a specific property in the given text.

    Args:
        text (str): The text to look for the property .
        prop (str): The property name to search for.
        value (str): The new value to replace the old value.

    Returns:
        str: The modified text with the replaced value, or None if the property is not found.
    """
    try:
        pattern = rf'^(input.*?;\s*//\s*{re.escape(prop)}\s*)$'
        line = re.findall(pattern, text, re.MULTILINE)
        
        if not line:
            logger.warning(f"Property '{prop}' not found in the text.")
            return None
        
        result = re.sub(pattern, replace_value(line[0], value), text, flags=re.MULTILINE)
        logger.info(f"Successfully replaced the value for property '{prop}'.")
        
        return result
    except Exception as e:
        logger.error(f"An error occurred in get_property_line: {e}")
        return None

def access_application(exe_path, timeout):
    """
    Start or connect to the application.

    Args:
        exe_path (str): The path to the executable file.
        timeout (int): The timeout period in seconds.

    Returns:
        Application: The connected or started application instance.
    """
    pid = is_application_running(exe_path)

    if pid:
        logger.info(f"Connecting to the running instance with PID: {pid}.")
        app = Application(backend="win32").connect(process=pid)
    else:
        logger.info(f"Starting the application from path: {exe_path}.")
        app = Application(backend="win32").start(exe_path, timeout=timeout)
    
    return app

def wait_for_window(app, title_re, timeout):
    """
    Wait for a window with a given title regular expression to exist.

    Args:
        app (Application): The application instance.
        title_re (str): The title regular expression to match the window.
        timeout (int): The timeout period in seconds.

    Returns:
        WindowSpecification: The matched window.
    """
    logger.info(f"Waiting for window with title matching: {title_re}.")
    try:
        window = app.window(title_re=title_re)
        window.wait("exists visible", timeout=timeout)
        logger.info("Window found and ready.")
        return window
    except Exception as e:
        logger.error(f"Exception occurred while waiting for the window: {e}")
        raise

def is_strategy_tester_open(window, timeout):
    """
    Check if the Strategy Tester window is open in the specified window.

    Args:
        window (WindowSpecification): The main window instance.
        timeout (int): The timeout period in seconds.

    Returns:
        tuple: (boolean, window) indicating if it's open and the window instance.
    """
    logger.info("Checking if the Strategy Tester is open...")
    try:
        strategy_tester = window.child_window(title="Tester")
        strategy_tester.wait('exists visible', timeout=timeout)
        logger.info("Strategy Tester is open.")
        return (True, strategy_tester)
    except Exception as e:
        logger.error(f"Exception occurred while checking the Strategy Tester: {e}")
        return (False, None)

def read_excel_to_dict(excel_path):
    """
    Reads an Excel file and converts each row into a dictionary, where the keys are the column headers
    and the values are the corresponding cell values in that row.

    Args:
        excel_path (str): The path to the Excel file.

    Returns:
        list of dict: A list of dictionaries, each representing a row in the Excel sheet.

    Raises:
        FileNotFoundError: If the Excel file is not found.
        ValueError: If the Excel sheet is empty or improperly formatted.
    """
    try:
        logger.info(f"Attempting to load the workbook from {excel_path}")
        # Load the workbook and select the active sheet
        workbook = load_workbook(excel_path)
        sheet = workbook.active

        # Get the column headers
        headers = [cell.value for cell in sheet[1]]
        if not headers:
            raise ValueError("No headers found in the Excel sheet")

        # Convert each row to a dictionary with column headers as keys and cell values as values
        settings_list = []
        for row in sheet.iter_rows(min_row=2, values_only=True):
            if any(row):  # Ensure the row is not entirely empty
                settings = {headers[i]: row[i] for i in range(len(headers))}
                settings_list.append(settings)

        if not settings_list:
            raise ValueError("No data rows found in the Excel sheet")

        logger.info("Successfully read the Excel file and converted it to a list of dictionaries")
        return settings_list

    except FileNotFoundError:
        logger.error(f"The file at {excel_path} was not found.")
        raise
    except ValueError as ve:
        logger.error(f"ValueError: {ve}")
        raise
    except Exception as e:
        logger.error(f"An unexpected error occurred: {e}")
        raise

def select_expert_advisor(strategy_tester):
    """
    Select 'Expert Advisor' from the combo box in the Strategy Tester.

    Args:
        strategy_tester (WindowSpecification): The Strategy Tester window.
    """
    try:
        logger.info("Checking for 'Indicator' combo box...")
        indicator_combo = strategy_tester.child_window(title="Indicator", class_name="ComboBox")
        indicator_combo.wait('exists visible', timeout=1)
        
        logger.info("Changing the value of 'Indicator' combo box to 'Expert Advisor'.")
        indicator_combo.select("Expert Advisor")
        logger.info("Successfully changed the value to 'Expert Advisor'.")
    except Exception as e:
        logger.error(f'It is likely that "Expert Advisor" has already been chosen')

def choose_EA(strategy_tester, ea_name):
    """
    Select the EA within the strategy tester.

    Args:
        strategy_tester (WindowSpecification): The Strategy Tester window.
        ea_name (str): The EA to choose.
    """
    try:
        
        logger.info("Searching for the Expert input in the Strategy Tester")
        combo_boxes = strategy_tester.children(class_name="ComboBox")
        ea_box_found = False

        for combo in combo_boxes:
            title = combo.window_text()
            if "ex4" in title: # If the combo box has ex4 in its text, it's an Expert input
                combo.click_input()  # Click to select the combo box
                ea_box_found = True
                i = 0
                prev_option = ''
                current_option = combo.window_text()
                
                logger.info(f"Navigating through options to find '{ea_name}'")

                # Use DOWN key to navigate through the dropdown options
                while i in range(50):
                    if ea_name == current_option: # if the EA is found
                        logger.info(f"Found '{ea_name}' in the options.")
                        return True
                    
                    send_keys('{DOWN}')
                    time.sleep(0.1)
                    prev_option = current_option
                    current_option = combo.window_text()
                    if prev_option == current_option: # if the limit of the EAs has been reached 
                        break

                    i += 1
                
                # Use UP key to navigate through the dropdown options
                i = 0
                while i in range(50):
                    if ea_name == current_option: # if the EA is found
                        logger.info(f"Found '{ea_name}' in the options.")
                        return True
                    
                    send_keys('{UP}')
                    time.sleep(0.1)
                    prev_option = current_option
                    current_option = combo.window_text()
                    if prev_option == current_option: # if the limit of the EAs has been reached 
                        return False

                    i += 1

        if not ea_box_found:
            logger.warning("No combo box with 'ex4' or 'ex5' found.")
            return False

    except Exception as e:
        logger.error(f"Exception occurred while selecting the combo box: {e}")
        return False

def choose_symbol(strategy_tester, symbol):
    """
    Select the symbol in the Strategy Tester.

    Args:
        strategy_tester (WindowSpecification): The Strategy Tester window.
        symbol (str): The symbol to choose.
    """
    try:
        logger.info("Searching for the Symbol input in the Strategy Tester")
        combo_boxes = strategy_tester.children(class_name="ComboBox")
        for combo in combo_boxes:
            title = combo.window_text()
            if "vs" in title or 'NOKSEK' in title: # If the Symbol input is found
                logger.info(f'Found symbol input: {combo}')
                if symbol in title: # If the wanted symbol has already been selected
                    logger.info(f'{symbol} has already been selected.')
                    return True
                else:
                    combo.click()
                    for i, full_symbol in enumerate(combo.item_texts()): # Select the wanted symbol
                        if symbol in full_symbol:
                            combo.select(i)
                            combo.set_keyboard_focus()
                            send_keys('{ENTER}')
                            logger.info(f'Found and selected {symbol} in the dropdown.')
                            return True

    except Exception as e:
        logger.error(f'Error has occurred when choosing the symbol: {e}')
        return False

def choose_period(strategy_tester, period):
    """
    Select the Period in the Strategy Tester.

    Args:
        strategy_tester (WindowSpecification): The Strategy Tester window.
        period (str): The period to choose.
    """
    try:
        logger.info("Searching for the Period input in the Strategy Tester")
        combo_boxes = strategy_tester.children(class_name="ComboBox")
        for combo in combo_boxes:
            title = combo.window_text()
            # If the Period input is found
            if "M1" in title or 'M5' in title or 'M15' in title or 'M30' in title or 'H1' in title or 'H4' in title or 'Daily' in title: 
                logger.info(f'Found Period input: {combo}')
                if period in title: # If the wanted period has already been selected
                    logger.info(f'{period} has already been selected.')
                    return True
                else:
                    combo.click()
                    for i, option in enumerate(combo.item_texts()): # Select the wanted period
                        if period in option:
                            combo.select(i)
                            combo.set_keyboard_focus()
                            send_keys('{ENTER}')
                            logger.info(f'Found and selected {period} in the dropdown.')
                            return True

    except Exception as e:
        logger.error(f'Error has occurred when choosing the symbol: {e}')
        return False

def choose_modelling(strategy_tester, model):
    """
    Select the Model in the Strategy Tester.

    Args:
        strategy_tester (WindowSpecification): The Strategy Tester window.
        model (str): The model to choose.
    """
    try:
        logger.info("Searching for the Model input in the Strategy Tester")
        combo_boxes = strategy_tester.children(class_name="ComboBox")
        for combo in combo_boxes:
            title = combo.window_text()
            # If the Model input is found
            if "Every tick" in title or 'Control points' in title or 'Open prices' in title: 
                logger.info(f'Found Model input: {combo}')
                if model in title: # If the wanted model has already been selected
                    logger.info(f'{model} has already been selected.')
                    return True
                else:
                    combo.click()
                    for i, option in enumerate(combo.item_texts()): # Select the wanted period
                        if model in option:
                            combo.select(i)
                            combo.set_keyboard_focus()
                            send_keys('{ENTER}')
                            logger.info(f'Found and selected {model} in the dropdown.')
                            return True

    except Exception as e:
        logger.error(f'Error has occurred when choosing the symbol: {e}')
        return False

def configure_visual_mode(strategy_tester):
    """
    Unchecks the Visual mode in the Strategy Tester (if it already isn't unchecked).

    Args:
        strategy_tester (WindowSpecification): The Strategy Tester window.
    """
    try:
        logger.info("Searching for the Visual mode button in the Strategy Tester")
        buttons = strategy_tester.children(class_name="Button")
        for btn in buttons:
            if btn.window_text() == 'Visual mode':
                btn.uncheck()
                logger.info("Visual mode unchecked.")
                return True
        
    except Exception as e:
        logger.error(f'Error has occurred when unchecking Visual mode: {e}')
        return False

def configure_dates(strategy_tester, from_date ,to_date):
    """
    This checks the Use date button and configures the From and To dates.

    Args:
        strategy_tester (WindowSpecification): The Strategy Tester window.
    """
    try:
        if from_date == None and to_date == None:
            logger.info("No dates to configure. Exiting.")
            return True

        logger.info("Searching for the Use date button in the Strategy Tester...")
        buttons = strategy_tester.children(class_name="Button")
        for btn in buttons:
            if btn.window_text() == 'Use date':
                btn.check()
                logger.info("Use date checked.")
                break

        logger.info("Searching for the dates in the Strategy Tester...")
        dates = strategy_tester.children(class_name="SysDateTimePick32")

        # Setting the From date
        if from_date is not None:
            _from = datetime.strptime(from_date, '%Y.%m.%d')
            dates[0].set_time(year=_from.year, month=_from.month, day=_from.day) 
        
        # Setting the To date
        if to_date is not None:
            to = datetime.strptime(to_date, '%Y.%m.%d')
            dates[1].set_time(year=to.year, month=to.month, day=to.day)

    except Exception as e:
        logger.error(f'Error has occurred when configuring the dates: {e}')
        return False

def configure_expert_properties(strategy_tester, ea_name, properties_string):
    """
    Configures the expert properties for a given Expert Advisor.

    Args:
        strategy_tester (pywinauto.Application): The application instance of the Strategy Tester.
        ea_name (str): The name of the Expert Advisor.
        properties_string (str): A comma-separated string of properties to set in the format
                                 'Property Name=Value'.
    """

    try:
        # If there are no properties to set, exit early
        if properties_string is None or properties_string.strip() == '':
            logger.info("No properties to set. Exiting.")
            return

        # Click on the Modify expert button and wait for the properties window to open
        strategy_tester.child_window(title="Modify expert", class_name="Button").click_input()
        logger.info("Clicked on 'Modify expert' button")

        # Connect to MetaEditor
        app = access_application(ME_EXE_PATH, TIMEOUT)
        editor = wait_for_window(app, title_re=".*MetaEditor.*", timeout=TIMEOUT)
        editor.set_focus()
        editor.maximize()

        # Go to the top of the file
        file = editor.child_window(best_match=ea_base_name(ea_name)+".mq4")
        file.set_focus()
        file.maximize()
        send_keys('^a^c')  # Select all content and copy it to the clipboard

        # Get the content from the clipboard
        original_code = paste()

        # Parse the properties string and set each property in the copied code
        modified_code = original_code
        for prop in properties_string.split(','):
            name, value = prop.split('=')
            name, value = name.strip(), value.strip()

            # Find where the variable is in the code and replace its value
            modified_code = change_input_value(modified_code, name, value)
            logger.info(f"Setting property '{name}' to '{value}'")

        # Copy the modified code back to the clipboard
        copy(modified_code)

        send_keys('^v')  # Select all content again and paste the modified content

        # Close MetaEditor
        send_keys('{F7}')  # Compile the file
        time.sleep(1.3) # Give the file time to compile
        send_keys('^{F4}')  # Close the current file
        send_keys('%{F4}')  # Close MetaEditor

        # Confirm the changes
        logger.info("Confirmed the changes and closed the Expert properties window")

    except Exception as e:
        logger.error(f"An error occurred while configuring expert properties: {e}")
        raise

def start_strategy_tester(strategy_tester):
    """
    Starts the Strategy Tester.

    Args:
        strategy_tester (WindowSpecification): The Strategy Tester window.
    """
    try:
        logger.info("Searching for the Start button in the Strategy Tester...")
        buttons = strategy_tester.children(class_name="Button")
        for btn in buttons:
            if btn.window_text() == 'Start':
                btn.click()
                logger.info("Clicked on 'Start' button")
                return True

        return False
    except Exception as e:
        logger.error(f"An error occurred while starting the Strategy Tester: {e}")
        raise

def main():
    try:
        # Start or connect to the application
        app = access_application(MT4_EXE_PATH, TIMEOUT)
        tradeview = wait_for_window(app, title_re=".*Tradeview.*", timeout=TIMEOUT)
        tradeview.set_focus()
        tradeview.maximize()

        # Check if the Strategy Tester is open. Open it if it's not
        tester_open, strategy_tester = is_strategy_tester_open(tradeview, 5)
        if not tester_open:
            logger.info("Strategy Tester is not open. Trying to open it.")
            tradeview.send_keystrokes('^r')  # Press Ctrl+R to open the Strategy Tester
            time.sleep(5)  # Wait for a few seconds to ensure the Strategy Tester has time to open
            tester_open, st = is_strategy_tester_open(tradeview, TIMEOUT)
            
            if not tester_open:
                logger.error("Failed to open the Strategy Tester. Exiting.")
                return

            strategy_tester = st

        # Get values from Excel sheet
        settings_list = read_excel_to_dict(EXCEL_PATH)
        for settings in settings_list:
            try:
                # Skip rows where 'Expert' is None
                if settings['Expert'] is None:
                    logger.info("Skipping row with missing 'Expert' value.")
                    continue

                # Make sure that "Expert Advisor" is selected in the Strategy Tester
                select_expert_advisor(strategy_tester)  

                # Select the EA
                ea_name = settings['Expert'].strip() 
                if not choose_EA(strategy_tester, ea_name):  
                    logger.error(f"Failed to select EA '{ea_name}'. Continuing.")
                    continue

                # Configure Expert properties
                configure_expert_properties(strategy_tester, settings['Expert'], settings['Expert properties'])
                tradeview = wait_for_window(app, title_re=".*Tradeview.*", timeout=10)
                
                # Select the symbol
                symbol = settings['Symbol'].strip()
                if not choose_symbol(strategy_tester, symbol):
                    logger.error(f"Failed to select symbol '{symbol}'. Continuing.")
                    continue

                # Select the period
                period = settings['Period'].strip()
                if not choose_period(strategy_tester, period):
                    logger.error(f"Failed to select period '{period}'. Continuing.")
                    continue
                
                # Select the model
                model = settings['Model'].strip()
                if not choose_modelling(strategy_tester, model):
                    logger.error(f"Failed to select model '{model}'. Continuing.")
                    continue

                # Configure Visual mode
                if not configure_visual_mode(strategy_tester):
                    logger.error(f"Failed to uncheck Visual mode. Continuing.")
                    continue

                # Configure the dates
                if not configure_dates(strategy_tester, settings['From'], settings['To']):
                    logger.error(f"Failed to configure the dates. Continuing.")
                    continue

                # Start the Strategy Tester
                if not start_strategy_tester(strategy_tester):
                    logger.error(f"Failed to start the Strategy Tester. Continuing.")
                    continue

                # Wait until the test finishes

            except Exception as e:
                logger.error(f"Exception occurred while configuring the Strategy Tester: {e}")
                logger.info('Continuing...')
                continue

    except Exception as e:
        logger.error(f"Exception occurred: {e}")

if __name__ == "__main__":
    main()
 
